"""
Step 3 — Extract real STP cost estimate data from the company's Excel files
into one clean, structured JSON file the pricing engine can use.

WHY THIS FILE EXISTS:
The raw Excel sheets are formatted for a human to read (title rows, blank
rows, a totals row mixed in with data rows). This script is the one-time
"translation" step: read the messy human spreadsheet, output clean data
a Python program can loop over without guessing at row positions.

This script does NOT do any pricing math. It only extracts and cleans.
The actual interpolation logic lives in interpolate.py (Step 4), on purpose,
so the two concerns (reading data vs. calculating prices) stay separate
and each is easy to test on its own.

INPUT:  raw_excel_source/STP <capacity> cum day/Estimate for STP & UF ....xlsx
OUTPUT: data/stp_boq_data.json

Run this manually whenever the source Excel files change:
    python pricing_engine/extract_excel_data.py
"""

import json
import re
from pathlib import Path

import openpyxl

# Folder this script lives in, so paths work no matter where you run it from
BASE_DIR = Path(__file__).resolve().parent
RAW_SOURCE_DIR = BASE_DIR / "raw_excel_source"
OUTPUT_PATH = BASE_DIR / "data" / "stp_boq_data.json"


def find_capacity_folders() -> dict[int, Path]:
    """
    Look inside raw_excel_source/ for folders named like "STP 200 cum day"
    and map capacity (as an int) -> folder path.

    We discover folders dynamically instead of hardcoding a list of
    capacities, so adding an 11th tier later just means adding a folder,
    not editing this script.
    """
    capacity_folders: dict[int, Path] = {}
    for folder in RAW_SOURCE_DIR.iterdir():
        if not folder.is_dir():
            continue
        match = re.search(r"(\d+)", folder.name)
        if match:
            capacity_folders[int(match.group(1))] = folder
    return dict(sorted(capacity_folders.items()))


def find_estimate_file(folder: Path) -> Path | None:
    """
    Find the .xlsx cost estimate file inside a capacity folder, if one
    exists. Some tiers (550, 600 cum/day) only have a Word quotation
    template and NO cost estimate — that's real, not a bug, so this
    function returns None for those rather than raising an error.
    """
    xlsx_files = list(folder.glob("*.xlsx"))
    if not xlsx_files:
        return None
    return xlsx_files[0]


def extract_line_items_from_sheet(filepath: Path) -> dict:
    """
    Read one capacity tier's estimate sheet and return its line items
    and total.

    The real sheet layout (confirmed by inspecting the actual files):
      - Row 6 is the header row (Sr. No., Equipments, Specification, Qty,
        Unit Rate, Amount)
      - Rows after that are line items, one per row, until a row with
        no Sr. No. but a value in the Amount column — that row is the
        grand total, not a line item.
    """
    workbook = openpyxl.load_workbook(filepath, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    line_items = []
    total = None

    for row in sheet.iter_rows(min_row=6, values_only=True):
        sr_no, name, spec, qty, unit_rate, amount = row[0], row[1], row[2], row[3], row[4], row[5]

        # The totals row: no Sr. No., no name, but an amount is present
        if sr_no is None and name is None and amount is not None:
            total = amount
            continue

        # A real line item: has a serial number and an equipment name
        if isinstance(sr_no, int) and name:
            line_items.append(
                {
                    "sr_no": sr_no,
                    "name": name.strip(),
                    "specification": spec.strip() if isinstance(spec, str) else spec,
                    "qty": qty,
                    "unit_rate": unit_rate,
                    "amount": amount,
                }
            )

    return {"total": total, "line_items": line_items}


def main() -> None:
    capacity_folders = find_capacity_folders()
    print(f"Found {len(capacity_folders)} capacity tier folders: {list(capacity_folders.keys())}")

    all_tiers = {}
    skipped_tiers = []

    for capacity, folder in capacity_folders.items():
        estimate_file = find_estimate_file(folder)

        if estimate_file is None:
            # Real situation, not an error: this tier has a quotation
            # template but no priced cost estimate yet (true for 550, 600).
            skipped_tiers.append(capacity)
            print(f"  {capacity} cum/day: NO cost estimate file found, skipping")
            continue

        tier_data = extract_line_items_from_sheet(estimate_file)
        all_tiers[capacity] = tier_data
        print(f"  {capacity} cum/day: extracted {len(tier_data['line_items'])} line items, total = {tier_data['total']}")

    output = {
        "source_note": (
            "Extracted from real company Excel cost estimates. "
            "Capacities present here have verified real pricing data. "
            "Tiers listed in 'unpriced_tiers' have a quotation template "
            "but no cost estimate yet, and must not be silently guessed."
        ),
        "unpriced_tiers": skipped_tiers,
        "tiers": all_tiers,
    }

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(output, indent=2))
    print(f"\nWrote clean data for {len(all_tiers)} tiers to {OUTPUT_PATH}")
    if skipped_tiers:
        print(f"Tiers with NO real pricing data (flagged, not guessed): {skipped_tiers}")


if __name__ == "__main__":
    main()